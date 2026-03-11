"""
report_generator.py
--------------------
Generates a polished Excel market intelligence report from the
Local Market Intelligence API response.

Usage:
    from report_generator import generate_excel_report
    import requests

    result = requests.post(
        "https://local-market-intelligence-api-1040656024374.us-central1.run.app/analyze",
        json={"business_type": "coffee shops", "location": "San Jose, CA"}
    ).json()

    output_path = generate_excel_report(result, output_path="market_report.xlsx")
"""

import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Style constants ────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN_BG   = "C6EFCE"
GREEN_FG   = "276221"
YELLOW_BG  = "FFEB9C"
YELLOW_FG  = "9C5700"
RED_BG     = "FCE4D6"
RED_FG     = "9C0006"
GOLD       = "FFD700"
SILVER     = "C0C0C0"
BRONZE     = "CD7F32"


def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)

def _font(bold=False, color="000000", size=10, italic=False):
    return Font(name="Arial", bold=bold, color=color, size=size, italic=italic)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


# ── Parsing helpers ────────────────────────────────────────────────────────────

def _parse_competitor_table(table_str: str) -> list[dict]:
    """Parse the markdown competitor table from the API response."""
    competitors = []
    if not table_str:
        return competitors
    lines = [l.strip() for l in table_str.strip().splitlines()]
    data_lines = [l for l in lines if l.startswith("|") and not set(l.replace("|", "").replace("-", "").replace(" ", "")) == set()]
    if len(data_lines) < 2:
        return competitors
    headers = [h.strip() for h in data_lines[0].split("|")[1:-1]]
    for line in data_lines[1:]:
        vals = [v.strip() for v in line.split("|")[1:-1]]
        if len(vals) == len(headers):
            row = dict(zip(headers, vals))
            competitors.append(row)
    return competitors


def _extract_metrics(report: dict) -> dict:
    """Pull summary_metrics or fall back to parsing the comprehensive_report."""
    metrics = report.get("summary_metrics", {})
    
    # Try to extract from comprehensive_report text if metrics are sparse
    comp = report.get("comprehensive_report", "")
    
    total = metrics.get("total_businesses_analyzed") or _regex_int(comp, r"(\d+)\s+direct competitors")
    avg_rating = _regex_float(comp, r"(\d+\.\d+)/5\.0 average rating") or \
                 _regex_float(comp, r"Average Rating.*?(\d+\.\d+)")
    competition = _regex_str(comp, r"Competition Level.*?[│|]\s*(\w+)")
    opportunities = metrics.get("opportunities_identified", 0)
    clusters = metrics.get("geographic_clusters", 0)
    sources = metrics.get("data_sources", ["Foursquare", "Google Places"])

    return {
        "total_competitors": total or "N/A",
        "avg_rating": avg_rating or "N/A",
        "competition_level": competition or "N/A",
        "opportunities": opportunities,
        "geographic_clusters": clusters,
        "data_sources": ", ".join(sources) if isinstance(sources, list) else sources,
    }


def _regex_int(text, pattern):
    m = re.search(pattern, text)
    return int(m.group(1)) if m else None

def _regex_float(text, pattern):
    m = re.search(pattern, text)
    return float(m.group(1)) if m else None

def _regex_str(text, pattern):
    m = re.search(pattern, text)
    return m.group(1).strip() if m else None


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _build_competitor_sheet(wb: Workbook, competitors: list[dict], business_type: str, location: str):
    ws = wb.active
    ws.title = "Competitor Data"

    # Title
    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = f"🏢 LOCAL MARKET INTELLIGENCE — {location} | {business_type.title()}"
    t.font = _font(bold=True, color=WHITE, size=13)
    t.fill = _fill(DARK_BLUE)
    t.alignment = _center()
    ws.row_dimensions[1].height = 32

    # Subtitle
    ws.merge_cells("A2:K2")
    s = ws["A2"]
    s.value = f"Competitor Analysis  |  Analyst-Verified  |  Generated: {datetime.now().strftime('%B %d, %Y')}"
    s.font = _font(italic=True, color="555555", size=9)
    s.fill = _fill("EAF0FB")
    s.alignment = _center()
    ws.row_dimensions[2].height = 18

    # Headers
    headers = ["Rank", "Business Name", "Rating", "# Reviews", "Price",
               "Category 1", "Category 2", "Source", "High Rated?", "Well Reviewed?", "Notes"]
    ws.row_dimensions[3].height = 28
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _font(bold=True, color=WHITE, size=11)
        c.fill = _fill(MID_BLUE)
        c.alignment = _center()
        c.border = _border()

    # Map API field names flexibly
    field_map = {
        "rank":     ["Rank", "rank"],
        "name":     ["Business Name", "business_name", "Name", "name"],
        "rating":   ["Rating", "rating"],
        "reviews":  ["Reviews", "# Reviews", "reviews", "review_count"],
        "price":    ["Price", "price", "Price Tier"],
        "cat1":     ["Categories", "categories", "Category", "Primary Category"],
        "source":   ["Source", "source", "Data Source"],
    }

    def get_field(row_dict, keys):
        for k in keys:
            if k in row_dict:
                return row_dict[k]
        return ""

    for i, row in enumerate(competitors):
        r = i + 4
        row_fill = _fill(WHITE) if i % 2 == 0 else _fill(LIGHT_GRAY)

        rank    = get_field(row, field_map["rank"]) or str(i + 1)
        name    = get_field(row, field_map["name"])
        rating  = get_field(row, field_map["rating"])
        reviews = get_field(row, field_map["reviews"])
        price   = get_field(row, field_map["price"])
        cats    = get_field(row, field_map["cat1"])
        source  = get_field(row, field_map["source"])

        # Clean up rating stars emoji if present
        rating_clean = str(rating).replace("⭐", "").strip()
        try:
            rating_val = float(rating_clean)
        except ValueError:
            rating_val = 0.0

        try:
            reviews_val = int(str(reviews).replace(",", ""))
        except ValueError:
            reviews_val = 0

        # Split categories if comma-separated
        cat_parts = [c.strip() for c in str(cats).split(",")]
        cat1 = cat_parts[0] if cat_parts else ""
        cat2 = cat_parts[1] if len(cat_parts) > 1 else ""

        high_rated    = "✓" if rating_val >= 8.0 else ""
        well_reviewed = "✓" if reviews_val >= 50 else ""

        vals = [rank, name, rating_clean, reviews_val, price, cat1, cat2, source,
                high_rated, well_reviewed, ""]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = _font(bold=(col == 2))
            c.fill = row_fill
            c.border = _border()
            c.alignment = _left() if col == 2 else _center()

        # Color-code rating
        rc = ws.cell(row=r, column=3)
        if rating_val >= 8.5:
            rc.fill = _fill(GREEN_BG)
            rc.font = _font(bold=True, color=GREEN_FG)
        elif rating_val >= 8.0:
            rc.fill = _fill(YELLOW_BG)
            rc.font = _font(bold=True, color=YELLOW_FG)
        elif rating_val > 0:
            rc.fill = _fill(RED_BG)
            rc.font = _font(color=RED_FG)

        ws.row_dimensions[r].height = 18

    col_widths = [6, 28, 9, 11, 8, 20, 20, 16, 11, 13, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"


def _build_summary_sheet(wb: Workbook, metrics: dict, opportunities_text: str,
                          business_type: str, location: str):
    ws = wb.create_sheet("Market Summary")

    ws.merge_cells("A1:F1")
    t = ws["A1"]
    t.value = f"📊 MARKET SUMMARY — {location} | {business_type.title()}"
    t.font = _font(bold=True, color=WHITE, size=13)
    t.fill = _fill(DARK_BLUE)
    t.alignment = _center()
    ws.row_dimensions[1].height = 32

    # Key metrics header
    ws.merge_cells("A3:F3")
    mh = ws["A3"]
    mh.value = "KEY MARKET METRICS"
    mh.font = _font(bold=True, color=WHITE, size=10)
    mh.fill = _fill(DARK_BLUE)
    mh.alignment = _center()
    ws.row_dimensions[3].height = 22

    metric_rows = [
        ("Total Competitors Identified",        str(metrics["total_competitors"])),
        ("Average Market Rating",               str(metrics["avg_rating"])),
        ("Competition Level",                   str(metrics["competition_level"])),
        ("Opportunities Identified",            str(metrics["opportunities"])),
        ("Geographic Clusters",                 str(metrics["geographic_clusters"])),
        ("Data Sources",                        metrics["data_sources"]),
        ("Analysis Date",                       datetime.now().strftime("%B %d, %Y")),
    ]

    for i, (label, value) in enumerate(metric_rows):
        r = i + 4
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _font(bold=True)
        lc.fill = _fill(LIGHT_BLUE if i % 2 == 0 else WHITE)
        lc.alignment = _left()
        lc.border = _border()
        ws.merge_cells(f"A{r}:C{r}")

        vc = ws.cell(row=r, column=4, value=value)
        vc.font = _font()
        vc.fill = _fill(LIGHT_BLUE if i % 2 == 0 else WHITE)
        vc.alignment = _center()
        vc.border = _border()
        ws.merge_cells(f"D{r}:F{r}")
        ws.row_dimensions[r].height = 18

    # Strategic opportunities
    opp_row = len(metric_rows) + 5
    ws.merge_cells(f"A{opp_row}:F{opp_row}")
    oh = ws[f"A{opp_row}"]
    oh.value = "🎯 STRATEGIC OPPORTUNITIES"
    oh.font = _font(bold=True, color=WHITE, size=10)
    oh.fill = _fill(MID_BLUE)
    oh.alignment = _center()
    ws.row_dimensions[opp_row].height = 22

    # Parse opportunities from text
    opps = _parse_opportunities(opportunities_text)
    if not opps:
        opps = [("No opportunities parsed", "Review the comprehensive report tab for details.")]

    for i, (title, desc) in enumerate(opps):
        r = opp_row + i + 1
        tc = ws.cell(row=r, column=1, value=title)
        tc.font = _font(bold=True, color=DARK_BLUE)
        tc.fill = _fill("EAF0FB" if i % 2 == 0 else WHITE)
        tc.alignment = _left()
        tc.border = _border()
        ws.merge_cells(f"A{r}:B{r}")

        dc = ws.cell(row=r, column=3, value=desc)
        dc.font = _font()
        dc.fill = _fill("EAF0FB" if i % 2 == 0 else WHITE)
        dc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        dc.border = _border()
        ws.merge_cells(f"C{r}:F{r}")
        ws.row_dimensions[r].height = 45

    for col, w in zip(["A","B","C","D","E","F"], [22, 18, 32, 14, 12, 12]):
        ws.column_dimensions[col].width = w


def _build_top5_sheet(wb: Workbook, competitors: list[dict], business_type: str, location: str):
    ws = wb.create_sheet("Top 5 Competitors")

    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = f"🏆 TOP 5 — {location} | {business_type.title()}"
    t.font = _font(bold=True, color=WHITE, size=13)
    t.fill = _fill(DARK_BLUE)
    t.alignment = _center()
    ws.row_dimensions[1].height = 32

    headers = ["", "Business Name", "Rating", "Reviews", "Price"]
    ws.row_dimensions[3].height = 26
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = _font(bold=True, color=WHITE, size=11)
        c.fill = _fill(MID_BLUE)
        c.alignment = _center()
        c.border = _border()

    medals = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]
    medal_fills = [GOLD, SILVER, BRONZE, LIGHT_BLUE, LIGHT_BLUE]

    top5 = competitors[:5]
    for i, row in enumerate(top5):
        r = i + 4
        name    = row.get("Business Name") or row.get("Name", "")
        rating  = str(row.get("Rating", "")).replace("⭐", "").strip()
        reviews = row.get("Reviews") or row.get("# Reviews", "")
        price   = row.get("Price") or row.get("Price Tier", "")

        vals = [medals[i], name, rating, reviews, price]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = _font(bold=(i < 3), size=11)
            c.fill = _fill(medal_fills[i])
            c.alignment = _left() if col == 2 else _center()
            c.border = _border()
        ws.row_dimensions[r].height = 24

    for col, w in zip(["A","B","C","D","E"], [6, 30, 10, 12, 10]):
        ws.column_dimensions[col].width = w


def _parse_opportunities(text: str) -> list[tuple[str, str]]:
    """Extract opportunity title/description pairs from the report text."""
    if not text:
        return []
    opps = []
    # Look for numbered or bulleted opportunity blocks
    blocks = re.split(r'\n(?=\d+\.|###|\*\*)', text)
    for block in blocks:
        lines = [l.strip() for l in block.strip().splitlines() if l.strip()]
        if not lines:
            continue
        title = re.sub(r'^[\d\.\*#\s]+', '', lines[0]).strip()
        desc  = " ".join(lines[1:])[:400] if len(lines) > 1 else ""
        if title:
            opps.append((title, desc))
    return opps[:5]  # cap at 5


# ── Public API ─────────────────────────────────────────────────────────────────

def generate_excel_report(api_response: dict, output_path: str = None) -> str:
    """
    Generate a polished Excel report from the API response.

    Parameters
    ----------
    api_response : dict
        The full JSON response from POST /analyze
    output_path : str, optional
        File path for the output .xlsx. Defaults to
        '{business_type}_{location}_MarketIntelligence.xlsx'

    Returns
    -------
    str
        Path to the saved Excel file.
    """
    analysis = api_response.get("analysis", api_response)

    # Extract business type and location from the report title if possible
    comp_report = analysis.get("comprehensive_report", "")
    biz_match   = re.search(r"##\s+(.*?)\s+Market Analysis\s+-\s+(.*?)\n", comp_report)
    business_type = biz_match.group(1) if biz_match else "Business"
    location      = biz_match.group(2) if biz_match else "Location"

    # Parse competitors
    competitor_table_str = analysis.get("competitor_table", "")
    competitors = _parse_competitor_table(competitor_table_str)

    # Extract metrics
    metrics = _extract_metrics(analysis)

    # Extract opportunities text
    opp_text = ""
    for section in ["strategic_opportunities", "opportunities", "comprehensive_report"]:
        if analysis.get(section):
            opp_text = analysis[section]
            break

    # Default output path
    if not output_path:
        safe_biz = re.sub(r'\W+', '_', business_type.title())
        safe_loc = re.sub(r'\W+', '_', location.replace(", ", "_"))
        output_path = f"{safe_biz}_{safe_loc}_MarketIntelligence.xlsx"

    wb = Workbook()

    _build_competitor_sheet(wb, competitors, business_type, location)
    _build_summary_sheet(wb, metrics, opp_text, business_type, location)
    if competitors:
        _build_top5_sheet(wb, competitors, business_type, location)

    wb.save(output_path)
    print(f"✅ Report saved: {output_path}")
    return output_path


# ── CLI convenience ────────────────────────────────────────────────────────────

if __name__ == "__main__":
    from .client import MarketIntelligenceClient

    business_type = input("Business type (e.g. coffee shops): ").strip()
    location      = input("Location (e.g. San Jose, CA): ").strip()
    context       = input("Additional context (optional): ").strip()

    print(f"\n⏳ Analyzing {business_type} in {location}...")
    client = MarketIntelligenceClient()
    result = client.analyze(business_type, location, context)

    generate_excel_report(result.raw)
